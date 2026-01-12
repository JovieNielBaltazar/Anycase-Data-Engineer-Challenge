"""
BIR Document Scraper - Unified Script
=====================================
Scrapes legal documents from 3 Bureau of Internal Revenue (Philippines) sources:

1. Legal Rulings - API endpoint (template 3322)
2. Revenue Delegation Authority Orders (RDAO) - API endpoint (template 3708)
3. PDF Documents - Direct URL extraction

Output: CSV files with structured document data

Author: [Your Name]
Date: January 2025
"""

import requests
import re
import os
from bs4 import BeautifulSoup
from datetime import datetime
from dataclasses import dataclass, asdict
from typing import List, Optional
import time

try:
    from openpyxl import Workbook
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


# =============================================================================
# CONFIGURATION
# =============================================================================

OUTPUT_DIR = r"C:\BIR Extracts"

API_BASE_URL = "https://bir-cms-ws.bir.gov.ph/api/pub/templates"
API_HEADERS = {
    "Accept": "*/*",
    "Client-Website-Id": "2",
    "Origin": "https://www.bir.gov.ph",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
}

# Source configurations
SOURCES = {
    "legal_rulings": {
        "template_id": 3322,
        "output_file": "legal_rulings.xlsx",
        "source_url": "https://www.bir.gov.ph/Legal-Matters-Legal-and-Legislative-Rulings"
    },
    "rdao": {
        "template_id": 3708,
        "output_file": "rdao_orders.xlsx",
        "source_url": "https://www.bir.gov.ph/2024-Revenue-Delegation-Authority-Orders"
    },
    "pdf_documents": {
        "urls": ["https://bir-cdn.bir.gov.ph/local/pdf/RA-10963-RRD.pdf"],
        "output_file": "pdf_documents.xlsx"
    }
}


# =============================================================================
# DATA MODEL
# =============================================================================

@dataclass
class Document:
    """Unified document model for all sources"""
    title: str = ""
    document_number: str = ""
    document_date: str = ""
    year: str = ""
    subject_matter: str = ""
    document_type: str = ""
    category: str = ""
    pdf_url: str = ""
    source_url: str = ""
    scraped_at: str = ""

    def to_dict(self) -> dict:
        return asdict(self)


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def fetch_api(template_id: int, max_retries: int = 3) -> Optional[dict]:
    """Fetch data from BIR API with retry logic"""
    url = f"{API_BASE_URL}/{template_id}/datasets"
    params = {"per_page": 3000}

    for attempt in range(max_retries):
        try:
            response = requests.get(
                url,
                params=params,
                headers=API_HEADERS,
                timeout=30,
                verify=True
            )
            response.raise_for_status()
            return response.json()

        except requests.exceptions.SSLError:
            print(f"    SSL Error (attempt {attempt + 1}/{max_retries})")
            if attempt == max_retries - 1:
                # Try with SSL verification disabled as fallback
                try:
                    import urllib3
                    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
                    response = requests.get(url, params=params, headers=API_HEADERS, timeout=30, verify=False)
                    response.raise_for_status()
                    return response.json()
                except Exception as e:
                    print(f"    Failed: {e}")
                    return None
            time.sleep(3)

        except requests.RequestException as e:
            print(f"    Error (attempt {attempt + 1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                time.sleep(3)

    return None


def normalize_document_number(text: str, doc_type: str) -> str:
    """
    Extract and normalize document number based on document type.

    Returns format:
        - BIR Rulings: "XXX-YYYY" (e.g., "001-2025")
        - RDAO: "XX-YYYY" (e.g., "01-2024")
        - PDF: Just numbers (e.g., "10963")
    """
    if not text:
        return ""

    if doc_type == "ruling":
        # Match patterns like "001-2025", "18-2025"
        match = re.search(r'(\d{1,3})[-–](\d{4})', text)
        if match:
            return f"{match.group(1).zfill(3)}-{match.group(2)}"

    elif doc_type == "rdao":
        # Match patterns like "35-2024"
        match = re.search(r'(\d{1,2})[-–](\d{4})', text)
        if match:
            return f"{match.group(1).zfill(2)}-{match.group(2)}"

    elif doc_type == "pdf":
        # Extract just numbers
        match = re.search(r'RA-(\d+)', text, re.IGNORECASE)
        if match:
            return match.group(1)
        match = re.search(r'(\d+)', text)
        if match:
            return match.group(1)

    return text


def parse_date(date_text: str) -> str:
    """Parse various date formats to ISO format (YYYY-MM-DD)"""
    if not date_text:
        return ""

    date_formats = [
        "%B %d, %Y",    # December 19, 2024
        "%B %d,%Y",     # December 19,2024
        "%b %d, %Y",    # Dec 19, 2024
        "%m/%d/%Y",     # 12/19/2024
    ]

    for fmt in date_formats:
        try:
            dt = datetime.strptime(date_text.strip(), fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue

    return date_text


def extract_year(text: str) -> str:
    """Extract 4-digit year from text"""
    match = re.search(r'(\d{4})', text)
    return match.group(1) if match else ""


def save_to_xlsx(documents: List[Document], output_path: str) -> bool:
    """Save documents to XLSX file with proper text formatting"""
    if not documents:
        print("  No documents to save")
        return False

    if not XLSX_AVAILABLE:
        print("  Error: openpyxl not installed. Run: pip install openpyxl")
        return False

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    fieldnames = [
        'title', 'document_number', 'document_date', 'year',
        'subject_matter', 'document_type', 'category',
        'pdf_url', 'source_url', 'scraped_at'
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"

    # Write header
    for col, header in enumerate(fieldnames, 1):
        ws.cell(row=1, column=col, value=header)

    # Write data - all as text to prevent date conversion
    for row_num, doc in enumerate(documents, 2):
        doc_dict = doc.to_dict()
        for col, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=row_num, column=col)
            cell.value = doc_dict.get(field, '')
            # Force text format for document_number column
            if field == 'document_number':
                cell.number_format = '@'  # Text format

    wb.save(output_path)
    print(f"  Saved {len(documents)} documents to {output_path}")
    return True


# =============================================================================
# SOURCE 1: LEGAL RULINGS
# =============================================================================

def scrape_legal_rulings() -> List[Document]:
    """Scrape BIR Legal Rulings from API"""
    print("\n[Source 1] Legal Rulings")
    print("-" * 40)

    config = SOURCES["legal_rulings"]
    documents = []

    print(f"  Fetching from template {config['template_id']}...")
    data = fetch_api(config["template_id"])

    if not data or "data" not in data:
        print("  Failed to fetch data")
        return documents

    print(f"  Found {len(data['data'])} year groups")

    for item in data["data"]:
        year_label = item.get("content", {}).get("Year", "")
        rulings_html = item.get("content", {}).get("Rulings", "")

        if not rulings_html:
            continue

        soup = BeautifulSoup(rulings_html, 'html.parser')
        links = soup.find_all('a', href=True)

        year = extract_year(year_label)
        count = 0

        for link in links:
            href = link.get('href', '')
            if not href.endswith('.pdf'):
                continue

            raw_title = link.get('title', '') or link.get_text(strip=True)
            doc_number = normalize_document_number(raw_title or href, "ruling")

            doc = Document(
                title=f"BIR Ruling No. {doc_number}" if doc_number else "",
                document_number=doc_number,
                document_date="",
                year=year,
                subject_matter="",
                document_type="BIR Ruling",
                category="Legal and Legislative Rulings",
                pdf_url=href,
                source_url=config["source_url"],
                scraped_at=datetime.now().isoformat()
            )
            documents.append(doc)
            count += 1

        print(f"    {year_label}: {count} rulings")

    return documents


# =============================================================================
# SOURCE 2: RDAO ORDERS
# =============================================================================

def scrape_rdao() -> List[Document]:
    """Scrape Revenue Delegation Authority Orders from API"""
    print("\n[Source 2] Revenue Delegation Authority Orders")
    print("-" * 40)

    config = SOURCES["rdao"]
    documents = []

    print(f"  Fetching from template {config['template_id']}...")
    data = fetch_api(config["template_id"])

    if not data or "data" not in data:
        print("  Failed to fetch data")
        return documents

    for item in data["data"]:
        content = item.get("content", {})

        # Find HTML table in content
        html_content = None
        if isinstance(content, str):
            html_content = content
        elif isinstance(content, dict):
            for value in content.values():
                if isinstance(value, str) and '<table' in value.lower():
                    html_content = value
                    break

        if not html_content:
            continue

        soup = BeautifulSoup(html_content, 'html.parser')
        rows = soup.find_all('tr')

        for row in rows:
            cells = row.find_all('td')
            if len(cells) < 3:
                continue

            # Extract cell data
            issuance_text = cells[0].get_text(strip=True)
            date_text = cells[2].get_text(strip=True)

            # Extract subject matter (without link text)
            subject_cell = BeautifulSoup(str(cells[1]), 'html.parser')
            for a_tag in subject_cell.find_all('a'):
                a_tag.decompose()
            subject = re.sub(r'\s+', ' ', subject_cell.get_text(strip=True)).strip()
            subject = re.sub(r'\s*\|\s*', ' ', subject).strip()

            # Get PDF URLs
            links = cells[1].find_all('a', href=True)
            full_text_url = ""
            for link in links:
                if 'full' in link.get('title', '').lower() or 'full' in link.get_text().lower():
                    full_text_url = link.get('href', '')
                    break
            if not full_text_url and links:
                full_text_url = links[-1].get('href', '')

            doc_number = normalize_document_number(issuance_text, "rdao")

            doc = Document(
                title=f"RDAO No. {doc_number}" if doc_number else "",
                document_number=doc_number,
                document_date=parse_date(date_text),
                year=extract_year(doc_number),
                subject_matter=subject,
                document_type="Revenue Delegation Authority Order",
                category="RDAO",
                pdf_url=full_text_url,
                source_url=config["source_url"],
                scraped_at=datetime.now().isoformat()
            )
            documents.append(doc)

    print(f"  Found {len(documents)} RDAOs")
    return documents


# =============================================================================
# SOURCE 3: PDF DOCUMENTS
# =============================================================================

def scrape_pdf_documents() -> List[Document]:
    """Extract metadata from PDF URLs"""
    print("\n[Source 3] PDF Documents")
    print("-" * 40)

    config = SOURCES["pdf_documents"]
    documents = []

    for url in config["urls"]:
        filename = url.split('/')[-1]
        title = filename.replace('.pdf', '')
        doc_number = normalize_document_number(filename, "pdf")

        doc = Document(
            title=title,
            document_number=doc_number,
            document_date="",
            year="",
            subject_matter="",
            document_type="",
            category="",
            pdf_url=url,
            source_url="",
            scraped_at=datetime.now().isoformat()
        )
        documents.append(doc)

        print(f"  {title}")
        print(f"    Document Number: {doc_number}")
        print(f"    PDF URL: {url}")

    return documents


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Main entry point"""
    print("=" * 60)
    print("BIR Document Scraper - Unified Script")
    print("=" * 60)
    print(f"Output Directory: {OUTPUT_DIR}")
    print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # Check if openpyxl is available
    if not XLSX_AVAILABLE:
        print("\nError: openpyxl is required. Install with: pip install openpyxl")
        return

    # Ensure output directory exists
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Source 1: Legal Rulings
    legal_rulings = scrape_legal_rulings()
    if legal_rulings:
        save_to_xlsx(legal_rulings, os.path.join(OUTPUT_DIR, SOURCES["legal_rulings"]["output_file"]))

    # Source 2: RDAO Orders
    rdao_orders = scrape_rdao()
    if rdao_orders:
        save_to_xlsx(rdao_orders, os.path.join(OUTPUT_DIR, SOURCES["rdao"]["output_file"]))

    # Source 3: PDF Documents
    pdf_docs = scrape_pdf_documents()
    if pdf_docs:
        save_to_xlsx(pdf_docs, os.path.join(OUTPUT_DIR, SOURCES["pdf_documents"]["output_file"]))

    # Combined output - all documents in one file
    all_documents = legal_rulings + rdao_orders + pdf_docs
    if all_documents:
        print("\n[Combined Output]")
        print("-" * 40)
        save_to_xlsx(all_documents, os.path.join(OUTPUT_DIR, "unified_data.xlsx"))

    # Summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"  Legal Rulings:  {len(legal_rulings)} documents")
    print(f"  RDAO Orders:    {len(rdao_orders)} documents")
    print(f"  PDF Documents:  {len(pdf_docs)} documents")
    print(f"  TOTAL:          {len(all_documents)} documents")
    print("-" * 60)
    print("Output Files:")
    print(f"  - {SOURCES['legal_rulings']['output_file']}")
    print(f"  - {SOURCES['rdao']['output_file']}")
    print(f"  - {SOURCES['pdf_documents']['output_file']}")
    print(f"  - unified_data.xlsx (combined)")
    print("=" * 60)
    print("Complete!")


if __name__ == "__main__":
    main()